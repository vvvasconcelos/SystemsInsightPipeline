import os
import pandas as pd
import numpy as np
import os
import datetime
import json
import warnings
import networkx as nx
from types import SimpleNamespace
from openpyxl import load_workbook

class Extract:
    def __init__(self, file_path):
        self.file_path = file_path
        self.variables = []
        self.var_to_type = {}
        self.adjacency_matrix = None
        self.interactions_matrix = None

    def extract_settings(self, double_factor_interventions_setting=None):
        """ Extract all settings based on the json and Kumu files
        """
        # Load the adjacency matrix from the KUMU file
        #variable_names, var_to_type_init, df_adj, interactions_matrix, intervention_variables, variable_of_interest, centrality = self.adjacency_matrix_from_kumu()  
        self.adjacency_matrix_from_kumu()  
    
        #return self.variables, self.var_to_type, self.df_adj, self.interactions_matrix, self.intervention_variables, self.variable_of_interest, self.centrality
        #if s.interaction_terms:
        if np.abs(self.interactions_matrix).sum() > 0:  # Interaction terms specified
            print("Solving an SDM with interaction terms.")
            print("By default, only single interventions will be simulated. To simulate interventions on two variables simultaneously, set s.double_factor_interventions = True")
            s = SimpleNamespace(**{"interaction_terms" : 1,
                                   "solve_analytically" : 0,
                                   "double_factor_interventions" : 0})  # Default to single-factor interventions
        else:
            print("No interaction terms specified so will solve linear SDM.")
            s = SimpleNamespace(**{"interaction_terms" : 0,
                                   "solve_analytically" : 1,
                                   "double_factor_interventions" : 0})
            #s.interaction_terms = False

        if double_factor_interventions_setting is not None: # If specified, override the default setting
            s.double_factor_interventions = double_factor_interventions_setting

        if s.double_factor_interventions and s.interaction_terms == False:
            warnings.warn("Without interaction terms, double factor interventions are not meaningful." \
                          "Consider setting double_factor_interventions to False.")

        # Load variable names and fill any spaces with underscores
        # s.stocks = [var.replace(" ", "_") for var in self.variables if self.var_to_type[var].lower() == 'stock']
        # s.auxiliaries = [var.replace(" ", "_") for var in self.variables if self.var_to_type[var].lower() == 'auxiliary']
        # s.constants = [var.replace(" ", "_") for var in self.variables if self.var_to_type[var].lower() == 'constant']
        # s.variables = [var.replace(" ", "_") for var in self.variables]  # s.auxiliaries + s.stocks + s.constants
        # s.stocks_and_constants = [var.replace(" ", "_") for var in self.variables if self.var_to_type[var] in ['stock', 'constant']]
        # s.stocks_and_auxiliaries = [var.replace(" ", "_") for var in self.variables if self.var_to_type[var] in ['stock', 'auxiliary']]
        # s.var_to_type = {var.replace(" ", "_") : self.var_to_type[var] for var in self.variables}

        s.stocks = [var for var in self.variables if self.var_to_type[var].lower() == 'stock']
        s.auxiliaries = [var for var in self.variables if self.var_to_type[var].lower() == 'auxiliary']
        s.constants = [var for var in self.variables if self.var_to_type[var].lower() == 'constant']
        s.variables = [var for var in self.variables]  # s.auxiliaries + s.stocks + s.constants
        s.stocks_and_constants = [var for var in self.variables if self.var_to_type[var] in ['stock', 'constant']]
        s.stocks_and_auxiliaries = [var for var in self.variables if self.var_to_type[var] in ['stock', 'auxiliary']]
        s.var_to_type = {var : self.var_to_type[var] for var in self.variables}

        s.variable_of_interest = []
        for voi in self.variable_of_interest:
            #voi = voi.replace(" ", "_")
            s.variable_of_interest += [voi]

        s.centrality = self.centrality
        s.intervention_variables = self.intervention_variables #[var.replace(" ", "_") for var in self.intervention_variables]
        s.intervention_strengths = self.intervention_strengths #{x.replace(" ", "_") : self.intervention_strengths[x] for x in self.intervention_strengths}

        if len(s.intervention_variables) == 0:
            raise(Exception("There should be at least one intervention variable specified in the Excel file."))
        
        # If double factor interventions selected, add double factor interventions 
        if s.double_factor_interventions:
            double_intervention_variables = []
            for i, var in enumerate(s.intervention_variables):
                for j in range(i + 1, len(s.intervention_variables)):
                    var_2 = s.intervention_variables[j]
                    double_intervention_variables += [var + '+' + var_2]
            
            s.intervention_variables += double_intervention_variables

        self.df_adj.rename(index=dict(zip(self.variables, s.variables)),
                           columns=dict(zip(self.variables, s.variables)), inplace=True)

        s.df_adj = self.df_adj  # Save the adjacency matrix to the settings
        s.interactions_matrix = self.interactions_matrix # Save the interactions matrix to the settings

        # Add the interactions to the adjacency matrix for the identification of feedback loops with interaction terms
        s.df_adj_incl_interactions = s.df_adj.copy()
        to_list, from1_list, from2_list = np.nonzero(s.interactions_matrix)
        for i in range(int(np.abs(s.interactions_matrix).sum())):
            to, from1, from2 = to_list[i], from1_list[i], from2_list[i]
            value = s.interactions_matrix[to, from1, from2]
            # Ensure that the interaction is nonzero in the adjacency matrix
            s.df_adj_incl_interactions.loc[s.df_adj_incl_interactions.index[to],
                                           s.df_adj_incl_interactions.columns[from1]] = value
            s.df_adj_incl_interactions.loc[s.df_adj_incl_interactions.index[to],
                                           s.df_adj_incl_interactions.columns[from2]] = value
            
        self.s = s  # Save the settings

        return s

    def check_loops(self, df_e, df_c):
        """ Check whether all loops have stocks and the ratio of balancing loops in the CLD
        """
        #stocks = list(df_e.loc[df_e.Type == "stock", "Label"])
        #num_stocks_and_auxiliaries = df_e.loc[(df_e.Type == "stock") | (df_e.Type == "auxiliary")].shape[0]
        stocks = [var for var in self.variables if self.var_to_type[var].lower() == "stock"]

        num_stocks_and_auxiliaries = len([var for var in self.variables 
                                        if self.var_to_type[var].lower() in ["stock", "auxiliary"]])
    
        # ## Check for balancing loops
        max_loops_check = 5  # Maximum number of loops to check to not take too much time
        if num_stocks_and_auxiliaries < max_loops_check:
            max_loops_check = num_stocks_and_auxiliaries

        # # Create a directed graph
        # G = nx.DiGraph()

        # # Add edges to the graph
        # for index, row in df_c.iterrows():
        #    G.add_edge(row['From'], row['To'])

        if (self.df_adj == -999).any().any() > 0:
        #    print("Cannot assess balancing vs. reinforcing feedback loops because polarities are missing")
            temp_df = self.df_adj.copy()
            temp_df = temp_df.replace(-999, 1)  # Replace missing polarities with 1, just cannot assess balancing loops
        
        G = nx.from_numpy_array(np.array(self.df_adj).T, create_using=nx.DiGraph)
        var_names = [var for var in self.df_adj.columns]  #.replace(" ", "_")
        G = nx.relabel_nodes(G, dict(enumerate(var_names)))
        feedback_loops = list(nx.simple_cycles(G, length_bound=max_loops_check))
        num_loops = len(feedback_loops)

        if num_loops > 0:
            print("\n" + str(num_loops), "feedback loops of maximum length", max_loops_check)

            ### Check if any loops have no stocks
            loops_wo_stocks = []

            for loop in feedback_loops:
                if sum([1 for x in loop if x in stocks]) == 0:
                    loops_wo_stocks += [loop]

            if len(loops_wo_stocks) > 0:
                print(len(loops_wo_stocks), "loops do not have a stock, which is",
                    round((len(loops_wo_stocks)/len(feedback_loops))*100, 5), "% of all loops")
                print("Loops without stocks:", loops_wo_stocks)
                raise(Exception("All loops should have at least one stock, redo the labeling"))
            else:
                print("All loops have at least one stock")
            
            ### Check the ratio of balancing loops
            if (self.df_adj == -999).any().any() > 0:
                num_balancing = 0
                balancing_loops = []
                for loop in feedback_loops:
                    num_min = 0
                    loop += [loop[0]]  # Add the first element to the end to close the loop
                    for i in range(len(loop)-1):
                        pol = df_c.loc[((df_c.From==loop[i])*1 + (df_c.To==loop[i+1])*1) == 2, "Type"].values[0]
                        if str(pol) == "-":
                            num_min += 1
                    if num_min % 2 != 0:  # If the number of negative polarities is odd
                    #    print(loop)
                        num_balancing += 1
                        balancing_loops += [loop]
                
                print(num_balancing, "(" + str(round((num_balancing/num_loops)*100, 2))+
                    "%) of these loops are balancing loops")
                # print(balancing_loops)
                if max_loops_check==num_stocks_and_auxiliaries:
                    print("The max length of loops checked is equal to the number of stocks and auxiliaries; all loops are considered\n")
                else:
                    print("The max length of loops checked is smaller than the number of stocks and auxiliaries; there may be more loops in the CLD\n")
        else:
            print("No feedback loops found in the CLD")

        # Calculate betweenness and closeness centrality
        bc = nx.betweenness_centrality(G, k=None, normalized=True)
        cc = nx.closeness_centrality(G.reverse(), wf_improved=False) # Apply to G.reverse() for outward distance vs inward
        bc = dict(sorted(bc.items(), key=lambda item: item[1], reverse=True))
        cc = dict(sorted(cc.items(), key=lambda item: item[1], reverse=True)) 
        #print('betweenness:')#, bc)
        #for key in bc:
        #    print(key, ":", round(bc[key], 3))
        #print('\ncloseness:')#, cc)  
        #for key in cc:  
        #    print(key, ":", round(cc[key], 3))

        self.centrality = {"betweenness" : bc, "closeness" : cc} #, "communicability" : nx.communicability(G)}

        
    def extract_adjacency_matrix(self):
        """Extract the adjacency matrix from an Excel table exported from Kumu (Kumu.io).
        The Kumu excel file contains one sheet with the CLD's variables ('Elements').
        It also contains a sheet with the CLD's causal links ('Connections').
        If there are known interactions in the system, these can be added in the 'Interactions' sheet.
        """
        # Read the elements, connections, and interactions sheets in the Kumu Excel file
        df_e = pd.read_excel(self.file_path, sheet_name="Elements")
        df_c = pd.read_excel(self.file_path, sheet_name="Connections")

        # Extract relevant columns
        df_e = df_e[["Label", "Type", "Tags", "Description"]]   
        df_c = df_c[["From", "Type", "To"]]
        
        # Extract variables from the Elements 
        self.original_variables = list(df_e["Label"])

        for var in self.original_variables:
            if "+" in var:
                raise(Exception(f'Variable name {var} contains a disallowed special character (+). Please remove these characters from the variable names.'))
            if "*" in var:
                raise(Exception(f'Variable name {var} contains a disallowed special character (*). Please remove these characters from the variable names.'))

        # Remove special characters from variable names and leave no extra spaces
        #self.variables =  [" ".join(var.replace("-", " ").replace("/", " ").replace("'", "").replace("(", "").replace(")", "").split()) for var in self.original_variables]
        #self.variables = [re.sub(r'[^a-zA-Z0-9_\s]', '', var).strip() for var in self.original_variables]
        #self.variables = [" ".join(var.split()) for var in self.variables]
        self.variables = [" ".join(var.split()) for var in self.original_variables] # Just remove extra spaces
        self.var_to_type = dict(zip(self.variables, list(df_e["Type"])))  # Create dictionary with relevant labels
        self.original_to_cleaned_var = dict(zip(self.original_variables, self.variables))

        ### Assign stock or constant to variables without type
        for var in self.original_variables:
            if str(self.var_to_type[self.original_to_cleaned_var[var]]).lower() not in ['stock', 'auxiliary', 'constant']:
                if var in list(df_c['To']): # If it has ingoing links and is therefore endogenous, assign as stock
                    print(f'Warning: Variable {self.original_to_cleaned_var[var]} has no (known) label assigned, namely {self.var_to_type[self.original_to_cleaned_var[var]]}, defaulting to stock as it has ingoing links.')
                    self.var_to_type[self.original_to_cleaned_var[var]] = 'stock'
                else:  # If exogenous, assign as constant
                    print(f'Warning: Variable {self.original_to_cleaned_var[var]} has no (known) label assigned, namely {self.var_to_type[self.original_to_cleaned_var[var]]}, defaulting to constant as it has no ingoing links.')
                    self.var_to_type[self.original_to_cleaned_var[var]] = 'constant'
                print("")

        self.intervention_variables = [self.original_to_cleaned_var[var] for var in list(df_e.loc[df_e["Tags"] != 0, "Label"])]
        self.intervention_strengths = dict(zip(self.variables, list(df_e.loc[:, "Tags"])))
        self.variable_of_interest = list(df_e.loc[df_e["Description"] == "VOI", "Label"])

        if len(self.variable_of_interest) == 1:
            #self.variable_of_interest = self.variable_of_interest[0]
            #if self.variable_of_interest != "A":
                print("Variable of interest:", self.variable_of_interest[0])
        else:
            #raise(Exception("There should be exactly one variable of interest in the CLD"))
                print("Variables of interest:", self.variable_of_interest)
        print("with", len(self.intervention_variables), "intervention variables")

        # Create an empty adjacency matrix
        num_variables = len(self.variables)
        # print("Num variables:", num_variables)
        self.adjacency_matrix = np.zeros((num_variables,
                                          num_variables))

        # Populate the adjacency matrix
        for i, origin in enumerate(df_c["From"]):
            if origin not in self.original_variables:
                raise(Exception(f'Origin variable {origin} in Connections sheet not found in Elements sheet. Please check the Kumu Excel file.'))
    
            destination = df_c["To"][i]

            if destination not in self.original_variables:
                raise(Exception(f'Destination variable {destination} in Connections sheet not found in Elements sheet. Please check the Kumu Excel file.'))

            # Determine the polarity
            polarity = 0
            temp = df_c["Type"][i]
            if str(temp) == '+':
                polarity = 1
            elif str(temp) == '-':
                polarity = -1
            else:  # No polarity specified
                polarity = -999

            # Calculate indices
            origin_index = self.original_variables.index(origin)
            destination_index = self.original_variables.index(destination)

            # Add polarity to adjacency matrix
            self.adjacency_matrix[destination_index, origin_index] = polarity

        # Create dataframe with adjacency matrix
        self.df_adj = pd.DataFrame(self.adjacency_matrix,
                                   columns=self.variables,
                                   index=self.variables) 

        #np.random.seed(s.seed)  # Set seed for reproducibility

        constants = [var for var in self.variables 
                     if self.var_to_type[var].lower() == 'constant']

        ### Check if any constants have incoming links
        for const in constants:
            num_incoming_links = np.sum(np.abs(self.df_adj.loc[const, :]))
            if num_incoming_links != 0:
                print(f'Removed {num_incoming_links} incoming links for constant {const}')
                self.df_adj.loc[const, :] = 0
                #else:
                #    raise(Exception(f'Number of incoming links for constant {const} is {num_links}, should be zero.'))

        for var in self.variables:
            if var not in constants:
                num_incoming_links = np.sum(np.abs(self.df_adj.loc[var, :]))
                #if num_incoming_links != 0:
                if num_incoming_links == 0:
                    #print(f'Removed {num_incoming_links} incoming links for variable {var}')
                    #self.df_adj.loc[var, :] = 0
                    raise(Exception(f'Number of incoming links for (non-constant) variable {var} is {num_incoming_links}, should be at least one.'))

        self.check_loops(df_e, df_c)  # Check for stocks in loops and ratio of balancing loops


    def extract_interactions_matrix(self):
        """Extract the interactions matrix from the 'Interactions' sheet in the Kumu Excel file."""
        wb = load_workbook(self.file_path, read_only=True)   # open an Excel file and return a workbook
        
        # Create an empty matrix to annotate interactions
        num_variables = len(self.variables)
        self.interactions_matrix = np.zeros((num_variables, num_variables, num_variables))

        if 'Interactions' in wb.sheetnames:
            df_i = pd.read_excel(self.file_path, sheet_name="Interactions")
            df_i = df_i[["From1", "From2", "Type", "To"]]

            # Create an empty matrix to annotate interactions
            #num_variables = len(self.variables)
            #self.interactions_matrix = np.zeros((num_variables, num_variables, num_variables))

            # Populate the interactions matrix
            for i, origin_1 in enumerate(df_i["From1"]):
                origin_2 = df_i["From2"][i]
                destination = df_i["To"][i]

                # Determine the polarity
                polarity = 0
                temp = df_i["Type"][i]
                if str(temp) == '+':
                    polarity = 1
                elif str(temp) == '-':
                    polarity = -1
                else:
                    polarity = -999 # Placeholder for missing values

                # Calculate indices
                origin_1_index = self.original_variables.index(origin_1)
                origin_2_index = self.original_variables.index(origin_2)
                destination_index = self.original_variables.index(destination)

                # Add polarity to interactions matrix
                self.interactions_matrix[destination_index, origin_2_index, origin_1_index] = polarity

    def adjacency_matrix_from_kumu(self):
        """Run the CLD analysis by extracting the adjacency matrix and interactions matrix."""
        self.extract_adjacency_matrix()
        self.extract_interactions_matrix()
        #return self.variables, self.var_to_type, self.df_adj, self.interactions_matrix, self.intervention_variables, self.variable_of_interest, self.centrality


### TESTING ###
    def test_extraction(self):
        """Test the CLD extraction by creating an examplar Kumu table and comparing the results."""
        # Create a sample evidence table
        data = {
            "From": ["A", "B", "C"],
            "Type": ["+", "-", "+"],
            "To": ["B", "C", "A"]
        }

        data_int = {
            "From1": ["A", "B"],
            "From2": ["C", "C"],
            "Type": ["+", "+"],
            "To": ["B", "A"]
        }

        df_e = pd.DataFrame(data["From"], columns=["Label"])
        df_e["Type"] = ["stock", "auxiliary", "constant"]
        df_e["Tags"] = ["VOI", -1, 1]
        df_c = pd.DataFrame(data)
        df_i = pd.DataFrame(data_int)

        # Save the evidence table to an Excel file
        original_file_path = self.file_path
        test_file_path = os.path.join(os.path.dirname(__file__), '..', 'test_files', 'evidence_table.xlsx')

        # Ensure the directory exists
        os.makedirs(os.path.dirname(test_file_path), exist_ok=True)

        with pd.ExcelWriter(test_file_path) as writer:
            df_e.to_excel(writer, sheet_name='Elements', index=False)
            df_c.to_excel(writer, sheet_name='Connections', index=False)
            df_i.to_excel(writer, sheet_name='Interactions', index=False)

        # Run the extraction
        self.file_path = test_file_path
        self.adjacency_matrix_from_kumu()
        self.file_path = original_file_path  # Set the original file path again
    
        # Define the expected results
        expected_adjacency_matrix = np.array([[0, 0, 1],
                                              [1, 0, 0],
                                              [0, -1, 0]])

        expected_interactions_matrix = np.array([[[0, 0, 0],
                                                  [0, 0, 0],
                                                  [0, 1, 0]], 
                                                 [[0, 0, 0],
                                                  [0, 0, 0],
                                                  [1, 0, 0]], 
                                                 [[0, 0, 0],
                                                  [0, 0, 0],
                                                  [0, 0, 0]]])

        # Assess the results
        assert np.all(expected_adjacency_matrix == self.adjacency_matrix)
        assert np.all(expected_interactions_matrix == self.interactions_matrix)
        assert np.all([x in self.variables for x in data["From"]])
        assert np.all([x in data["From"] for x in self.variables])
        print("Test for loading KUMU table passed.")
    